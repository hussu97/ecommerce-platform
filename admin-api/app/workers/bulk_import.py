"""
Bulk product import worker. Polls for pending uploads and processes them.
Run: cd admin-api && python -m app.workers.bulk_import
"""
from __future__ import annotations

import asyncio
import csv
import io
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

# Ensure admin-api is on path when run from project root
if str(Path(__file__).resolve().parents[2]) not in sys.path:
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.db.session import AsyncSessionLocal
from app.models.product import Product
from app.models.product_child import ProductChild
from app.models.product_translation import ProductTranslation
from app.models.product_attribute_value import ProductAttributeValue
from app.models.product_bulk_upload import ProductBulkUpload
from app.models.user import User
from app.models.taxonomy import Taxonomy
from app.models.taxonomy_attribute import TaxonomyAttribute, TaxonomyAttributeOption
from app.models.brand import Brand
from app.models.language import Language
from app.storage import get_storage
from app.core.codegen import generate_parent_code, generate_child_code
from app.core.config import settings


POLL_INTERVAL = 5


def _slugify(s: str) -> str:
    s = (s or "").lower().strip()
    s = re.sub(r"[^\w\s-]", "", s)
    s = re.sub(r"[\s_-]+", "-", s)
    return s.strip("-") or "untitled"


def _parse_xlsx(content: bytes) -> tuple[list[str], list[dict]]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(h or "").strip() for h in rows[0]]
    data = []
    for row in rows[1:]:
        row_dict = {}
        for i, h in enumerate(headers):
            if i < len(row) and row[i] is not None:
                val = row[i]
                if isinstance(val, (int, float)) and "price" in h.lower() or "stock" in h.lower() or "id" in h.lower():
                    row_dict[h] = val
                else:
                    row_dict[h] = str(val).strip() if val else ""
            else:
                row_dict[h] = ""
        data.append(row_dict)
    wb.close()
    return headers, data


def _parse_csv(content: bytes) -> tuple[list[str], list[dict]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    headers = reader.fieldnames or []
    data = list(reader)
    return list(headers), data


def _parse_tsv(content: bytes) -> tuple[list[str], list[dict]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text), delimiter="\t")
    headers = reader.fieldnames or []
    data = list(reader)
    return list(headers), data


def _parse_file(content: bytes, ext: str) -> tuple[list[str], list[dict]]:
    if ext == ".xlsx":
        return _parse_xlsx(content)
    if ext == ".csv":
        return _parse_csv(content)
    if ext == ".tsv":
        return _parse_tsv(content)
    raise ValueError(f"Unsupported format: {ext}")


async def _process_row(
    db: AsyncSession,
    row: dict,
    row_num: int,
    default_taxonomy_id: int,
    attr_name_to_option_ids: dict[str, dict[str, int]],
    lang_en_id: Optional[int],
    lang_ar_id: Optional[int],
    taxonomies_by_name: dict[str, int],
    taxonomies_by_slug: dict[str, int],
    brands_by_name: dict[str, Brand],
    brands_by_slug: dict[str, Brand],
    products_by_code: dict[str, Product],
) -> tuple[bool, Optional[str]]:
    """Process one row. Option A: single-size (is_single_size or no size_value + barcode) or multi-size (size_value).
    Every product gets at least one child."""
    parent_code_val = (row.get("parent_code") or "").strip()
    size_value_val = (row.get("size_value") or "").strip()
    barcode_val = (row.get("barcode") or "").strip()
    is_single = str(row.get("is_single_size") or "").strip().lower() in ("1", "true", "yes")

    # Adding child to existing parent (multi-size extra child)
    if parent_code_val and size_value_val:
        product = products_by_code.get(parent_code_val.upper()) or products_by_code.get(parent_code_val)
        if not product:
            return False, f"parent_code not found: {parent_code_val}"
        existing = await db.execute(
            select(ProductChild).where(
                ProductChild.product_id == product.id,
                ProductChild.size_value == size_value_val,
            )
        )
        if existing.scalar_one_or_none():
            return False, f"duplicate size_value {size_value_val} for product {product.code}"
        try:
            stock = int(float(row.get("stock_quantity") or 0))
        except (ValueError, TypeError):
            stock = 0
        child_code = await generate_child_code(db, product.id)
        child = ProductChild(
            product_id=product.id,
            code=child_code,
            barcode=barcode_val or None,
            size_value=size_value_val,
            stock_quantity=max(0, stock),
        )
        db.add(child)
        await db.flush()
        return True, None

    # New product (single-size or first row of multi-size)
    name_en = (row.get("name_en") or "").strip()
    name_ar = (row.get("name_ar") or "").strip()
    name = name_en or name_ar
    if not name:
        return False, "name_en or name_ar required"
    try:
        price = float(row.get("price") or 0)
    except (ValueError, TypeError):
        return False, "invalid price"
    if price < 0:
        return False, "price must be >= 0"
    stock = 0
    try:
        stock = int(float(row.get("stock_quantity") or 0))
    except (ValueError, TypeError):
        pass
    category_id = default_taxonomy_id
    cat_val = (row.get("category_name") or row.get("category_slug") or "").strip()
    if cat_val:
        t = taxonomies_by_name.get(cat_val.lower()) or taxonomies_by_slug.get(_slugify(cat_val))
        if t is not None:
            category_id = t
        else:
            return False, f"category not found: {cat_val}"
    brand_id = None
    brand_val = (row.get("brand_name") or row.get("brand_slug") or "").strip()
    if brand_val:
        b = brands_by_name.get(brand_val.lower()) or brands_by_slug.get(_slugify(brand_val))
        brand_id = b.id if b else None
        if b is None:
            return False, f"brand not found: {brand_val}"
    is_active = True
    av = str(row.get("is_active", "")).strip().lower()
    if av in ("0", "false", "no", "inactive"):
        is_active = False
    base_slug = _slugify(name)
    slug = base_slug
    n = 1
    while True:
        r = await db.execute(select(Product.id).where(Product.slug == slug))
        if r.scalar_one_or_none() is None:
            break
        slug = f"{base_slug}-{n}"
        n += 1
    code = await generate_parent_code(db)
    product = Product(
        code=code,
        name=name,
        slug=slug,
        description=(row.get("description_en") or row.get("description_ar") or "").strip() or None,
        price=price,
        stock_quantity=0,
        stock_reserved=0,
        image_url=(row.get("image_url") or "").strip() or None,
        category_id=category_id,
        brand_id=brand_id,
        is_active=is_active,
    )
    db.add(product)
    await db.flush()
    products_by_code[code] = product

    if name_en and lang_en_id:
        db.add(ProductTranslation(product_id=product.id, language_id=lang_en_id, name=name_en, description=(row.get("description_en") or "").strip() or None))
    if name_ar and lang_ar_id:
        db.add(ProductTranslation(product_id=product.id, language_id=lang_ar_id, name=name_ar, description=(row.get("description_ar") or "").strip() or None))
    if not name_en and not name_ar and lang_en_id:
        db.add(ProductTranslation(product_id=product.id, language_id=lang_en_id, name=name, description=(row.get("description_en") or row.get("description_ar") or "").strip() or None))
    for attr_name, value_to_opt_id in attr_name_to_option_ids.items():
        val = (row.get(attr_name) or "").strip()
        if not val:
            continue
        opt_id = value_to_opt_id.get(val.lower()) or value_to_opt_id.get(val)
        if opt_id:
            db.add(ProductAttributeValue(product_id=product.id, option_id=opt_id))

    # Single-size: one child with size_value single_size
    if is_single or (not size_value_val and (barcode_val or stock is not None)):
        child_code = await generate_child_code(db, product.id)
        child = ProductChild(
            product_id=product.id,
            code=child_code,
            barcode=barcode_val or None,
            size_value=settings.SINGLE_SIZE_VALUE,
            stock_quantity=max(0, stock),
        )
        db.add(child)
        await db.flush()
        return True, None

    # Multi-size first child
    if size_value_val:
        child_code = await generate_child_code(db, product.id)
        child = ProductChild(
            product_id=product.id,
            code=child_code,
            barcode=barcode_val or None,
            size_value=size_value_val,
            stock_quantity=max(0, stock),
        )
        db.add(child)
        await db.flush()
        return True, None

    # No size and not single: treat as single-size with no barcode
    child_code = await generate_child_code(db, product.id)
    child = ProductChild(
        product_id=product.id,
        code=child_code,
        barcode=None,
        size_value=settings.SINGLE_SIZE_VALUE,
        stock_quantity=max(0, stock),
    )
    db.add(child)
    await db.flush()
    return True, None


async def process_upload(upload: ProductBulkUpload, content: bytes, ext: str) -> None:
    async with AsyncSessionLocal() as db:
        r = await db.execute(select(ProductBulkUpload).where(ProductBulkUpload.id == upload.id))
        u = r.scalar_one_or_none()
        if u:
            u.status = "processing"
            u.started_at = datetime.utcnow()
            await db.commit()
    try:
        headers, rows = _parse_file(content, ext)
        total = len(rows)
    except Exception as e:
        async with AsyncSessionLocal() as db:
            r = await db.execute(select(ProductBulkUpload).where(ProductBulkUpload.id == upload.id))
            u = r.scalar_one_or_none()
            if u:
                u.status = "failed"
                u.error_details = {"0": str(e)}
                u.completed_at = datetime.utcnow()
            await db.commit()
        return
    async with AsyncSessionLocal() as db:
        r = await db.execute(select(ProductBulkUpload).where(ProductBulkUpload.id == upload.id))
        u = r.scalar_one_or_none()
        if not u:
            return
        u.total_rows = total
        await db.commit()
    async with AsyncSessionLocal() as db:
        result = await db.execute(
            select(Taxonomy)
            .where(Taxonomy.id == upload.taxonomy_id)
            .options(selectinload(Taxonomy.attributes).selectinload(TaxonomyAttribute.options))
        )
        taxonomy = result.scalar_one_or_none()
    if not taxonomy:
        async with AsyncSessionLocal() as db:
            r = await db.execute(select(ProductBulkUpload).where(ProductBulkUpload.id == upload.id))
            u = r.scalar_one_or_none()
            if u:
                u.status = "failed"
                u.error_details = {"0": "Taxonomy not found"}
                u.completed_at = datetime.utcnow()
            await db.commit()
        return
    attr_name_to_option_ids: dict[str, dict[str, int]] = {}
    for attr in taxonomy.attributes or []:
        if not attr.is_active:
            continue
        value_to_opt_id = {}
        for opt in attr.options or []:
            if opt.is_active:
                value_to_opt_id[opt.value.lower()] = opt.id
                value_to_opt_id[opt.value] = opt.id
        attr_name_to_option_ids[attr.name] = value_to_opt_id
    async with AsyncSessionLocal() as db:
        result = await db.execute(select(Language))
        langs = {l.code.lower(): l for l in result.scalars().all()}
        lang_en = langs.get("en")
        lang_ar = langs.get("ar")
        lang_en_id = lang_en.id if lang_en else None
        lang_ar_id = lang_ar.id if lang_ar else None
        result = await db.execute(select(Brand))
        brands = result.scalars().all()
        result = await db.execute(select(Taxonomy))
        taxonomies = result.scalars().all()
    brands_by_name = {b.name.lower(): b for b in brands}
    brands_by_slug = {(b.slug or "").lower(): b for b in brands if b.slug}
    taxonomies_by_name = {t.name.lower(): t.id for t in taxonomies}
    taxonomies_by_slug = {(t.slug or "").lower(): t.id for t in taxonomies if t.slug}
    processed = 0
    errors = 0
    error_details: dict = {}
    products_by_code: dict[str, Product] = {}
    async with AsyncSessionLocal() as db:
        for i, row in enumerate(rows):
            row_num = i + 2
            ok, err = await _process_row(
                db, row, row_num, upload.taxonomy_id or 0,
                attr_name_to_option_ids, lang_en_id, lang_ar_id,
                taxonomies_by_name, taxonomies_by_slug,
                brands_by_name, brands_by_slug,
                products_by_code,
            )
            if ok:
                processed += 1
                await db.commit()
            else:
                errors += 1
                error_details[str(row_num)] = err or "Unknown error"
                await db.rollback()
        r = await db.execute(select(ProductBulkUpload).where(ProductBulkUpload.id == upload.id))
        u = r.scalar_one_or_none()
        if u:
            u.status = "completed"
            u.processed_rows = processed
            u.error_rows = errors
            u.error_details = error_details if error_details else None
            u.completed_at = datetime.utcnow()
        await db.commit()


async def run_worker():
    storage = get_storage()
    print("Bulk import worker started. Polling for pending uploads...")
    while True:
        try:
            async with AsyncSessionLocal() as db:
                result = await db.execute(
                    select(ProductBulkUpload)
                    .where(ProductBulkUpload.status == "pending")
                    .order_by(ProductBulkUpload.created_at)
                    .limit(1)
                )
                upload = result.scalar_one_or_none()
                if upload:
                    try:
                        content = storage.get(upload.storage_path)
                    except Exception as e:
                        upload.status = "failed"
                        upload.error_details = {"0": f"Failed to read file: {e}"}
                        upload.completed_at = datetime.utcnow()
                        await db.commit()
                    else:
                        ext = os.path.splitext(upload.filename or "")[1].lower()
                        if ext not in {".xlsx", ".csv", ".tsv"}:
                            upload.status = "failed"
                            upload.error_details = {"0": f"Unsupported format: {ext}"}
                            upload.completed_at = datetime.utcnow()
                            await db.commit()
                        else:
                            await db.commit()
                            await process_upload(upload, content, ext)
        except Exception as e:
            print(f"Worker error: {e}", flush=True)
        await asyncio.sleep(POLL_INTERVAL)


def main():
    asyncio.run(run_worker())


if __name__ == "__main__":
    main()
