"""Generate bulk product import templates (xlsx, csv, tsv) based on taxonomy."""
import csv
import io
from typing import List, Tuple

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

from app.models.taxonomy import Taxonomy
from app.models.taxonomy_attribute import TaxonomyAttribute, TaxonomyAttributeOption


def _get_headers_and_attr_options(taxonomy: Taxonomy) -> Tuple[List[str], dict]:
    """Get column headers and attribute name -> option values for dropdowns.
    Uses names/slugs only (no IDs). category_name defaults to template taxonomy if empty."""
    headers = [
        "name_en", "name_ar", "description_en", "description_ar",
        "price", "stock_quantity", "image_url",
        "category_name", "brand_name",
        "is_active",
        "parent_code",  # optional: link to existing product (by code) when adding children only
        "size_value",   # optional: S, M, L, etc. Empty + barcode = single-size product (one child with single_size)
        "barcode",      # optional: per child
        "is_single_size",  # optional: 1/true = create one child with size_value single_size (barcode, stock_quantity used)
    ]
    attr_options: dict[str, List[str]] = {}
    if taxonomy.attributes:
        for attr in sorted(taxonomy.attributes, key=lambda a: (a.sort_order or 0, a.name)):
            if not attr.is_active:
                continue
            opt_values = [
                o.value for o in (attr.options or [])
                if o.is_active
            ]
            if opt_values:
                attr_options[attr.name] = opt_values
            headers.append(attr.name)
    return headers, attr_options


def generate_xlsx(taxonomy: Taxonomy) -> bytes:
    """Generate xlsx template with data validation dropdowns for attribute columns."""
    headers, attr_options = _get_headers_and_attr_options(taxonomy)
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    from openpyxl.utils import get_column_letter
    for col, h in enumerate(headers, start=1):
        if h in attr_options:
            opts = attr_options[h]
            opts_str = ",".join(str(v) for v in opts)
            dv = DataValidation(
                type="list",
                formula1=f'"{opts_str}"',
                allow_blank=True,
            )
            dv.error = "Value must be one of the options"
            dv.errorTitle = "Invalid value"
            ws.add_data_validation(dv)
            col_letter = get_column_letter(col)
            dv.add(f"{col_letter}2:{col_letter}10000")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_csv(taxonomy: Taxonomy) -> bytes:
    """Generate csv template (UTF-8 with BOM for Excel)."""
    headers, _ = _get_headers_and_attr_options(taxonomy)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    content = buf.getvalue()
    return ("\ufeff" + content).encode("utf-8")


def generate_tsv(taxonomy: Taxonomy) -> bytes:
    """Generate tsv template (UTF-8 with BOM for Excel)."""
    headers, _ = _get_headers_and_attr_options(taxonomy)
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter="\t")
    writer.writerow(headers)
    content = buf.getvalue()
    return ("\ufeff" + content).encode("utf-8")
